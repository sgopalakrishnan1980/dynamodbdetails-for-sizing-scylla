#!/usr/bin/env python3
"""
DynamoDB Logs to Excel Processor
================================

This script processes DynamoDB metrics log files and creates professional Excel files
with comprehensive analysis of DynamoDB performance metrics.

OVERVIEW:
---------
The script processes log files collected by the DynamoDB metrics collection tools
(Go/Bash versions) and converts them into Excel format for analysis and reporting.

KEY FEATURES:
-------------
1. Multi-Region Processing: Creates separate Excel files for each AWS region
2. Dynamic Operation Detection: Automatically discovers all DynamoDB operations
3. Comprehensive Data Extraction: Processes both CSV and JSON format log files
4. Table Metadata Integration: Extracts and includes table details from table_detailed.log
5. Professional Excel Output: Formatted worksheets with headers, styling, and auto-adjusted columns
6. Missing Data Logging: Detailed logging of APIs with no data for troubleshooting
7. Flexible Output: Customizable customer names and output directories

DIRECTORY STRUCTURE EXPECTED:
----------------------------
log_directory/
├── table_detailed.log                    # Table metadata in ASCII format
├── ap-northeast-1/                      # Region directory
│   └── table_name/                      # DynamoDB table name
│       └── operation/                   # DynamoDB operation (GetItem, Query, etc.)
│           ├── p99_latency/             # P99 latency metric files
│           │   └── *.log               # Individual log files with timestamps
│           └── sample_count/            # Sample count metric files
│               └── *.log               # Individual log files with timestamps
├── us-east-1/
└── ap-south-1/

DATA FORMATS SUPPORTED:
----------------------
1. CSV Format (Individual Log Files):
   DATAPOINTS <value> <timestamp> <unit>
   Example: DATAPOINTS 2935.0 2025-08-05T15:29:00+00:00 Milliseconds

2. JSON Format (CloudWatch API Response):
   {
     "Datapoints": [
       {
         "Timestamp": "2025-08-05T15:29:00Z",
         "SampleCount": 2935.0
       }
     ]
   }

3. ASCII Table Format (table_detailed.log):
   ||  TableSizeBytes           |  1156168014599                        ||
   ||  ItemCount                |  1670604447                           ||
   ||  BillingMode              |  PAY_PER_REQUEST                      ||

EXCEL OUTPUT STRUCTURE:
----------------------
For each region, creates: customer_region_dynamodb_metrics_YYYYMMDD_HHMMSS.xlsx

Each Excel file contains:
1. Table Summary Worksheet:
   - Table metadata (size, items, billing mode, etc.)
   - Max sample counts sorted by value
   - Max P99 latencies sorted by value

2. Individual Table Worksheets (one per DynamoDB table):
   - Sample Count data for all operations (rows 3-9)
   - P99 Latency data for all operations (rows 12-18)
   - Max values section with highlighting
   - Professional formatting and styling

USAGE:
------
python process_dynamodb_logs.py -t <source_directory> -d <destination_directory> -c <customer_name> -f <format>

ARGUMENTS:
----------
-t, --target: Source directory containing DynamoDB log files
-d, --destination: Destination directory for output Excel files
-c, --customer: Customer name prefix for Excel files
-f, --format: Data format ('json' or 'csv')

EXAMPLES:
---------
# Process CSV format logs
python process_dynamodb_logs.py -t /path/to/logs -d /tmp/output -c sas-2 -f csv

# Process JSON format logs
python process_dynamodb_logs.py -t /path/to/logs -d /tmp/output -c sas-2 -f json

OUTPUT FILES:
-------------
1. Excel files: customer_region_dynamodb_metrics_YYYYMMDD_HHMMSS.xlsx
2. Log file: customer_YYYYMMDD_HHMMSS.log

PROCESSING LOGIC:
-----------------
1. Region Discovery: Automatically finds all regions in the log directory
2. File Detection: Locates individual log files in the correct directory structure
3. Data Parsing: Extracts timestamps and values from log files
4. Metadata Extraction: Parses table details from ASCII table format
5. Excel Generation: Creates formatted Excel files with comprehensive data
6. Logging: Records all processing steps and missing data for troubleshooting

ERROR HANDLING:
--------------
- Validates input directories and parameters
- Gracefully handles missing files or directories
- Continues processing other regions if one fails
- Logs all errors and warnings for troubleshooting
- Provides detailed error messages for debugging

PERFORMANCE CONSIDERATIONS:
--------------------------
- Sequential processing of regions to avoid memory issues
- Efficient file I/O with minimal operations
- Error isolation to prevent one failure from affecting others
- Comprehensive logging for monitoring and debugging

AUTHOR: DynamoDB Metrics Collection Team
VERSION: 2.0
DATE: August 2025
"""

import os
import sys
import re
import glob
import json
import argparse
from datetime import datetime
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def detect_file_format(file_path):
    """
    Detect if a file contains JSON or CSV format data.
    
    This function analyzes the content of a log file to determine its format.
    It first attempts to parse as JSON, then checks for CSV patterns.
    
    CSV DETECTION LOGIC:
    - Looks for comma-separated values in the first line
    - Checks for tab-delimited data
    - Checks for semicolon-delimited data
    - Falls back to 'unknown' if no clear pattern is found
    
    JSON DETECTION LOGIC:
    - Attempts to parse the entire file as JSON
    - Returns 'json' if successful parsing
    - Falls back to CSV detection if JSON parsing fails
    
    Args:
        file_path (str): Path to the file to analyze
        
    Returns:
        str: 'json', 'csv', or 'unknown'
        
    Example:
        >>> detect_file_format('sample_count.log')
        'csv'
        >>> detect_file_format('cloudwatch_response.json')
        'json'
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read().strip()
            
        if not content:
            return 'unknown'
        
        # Try to parse as JSON first
        try:
            json.loads(content)
            return 'json'
        except (json.JSONDecodeError, ValueError):
            pass
        
        # Check if it looks like CSV
        lines = content.split('\n')
        if len(lines) > 0:
            # Check if first line contains commas (typical CSV delimiter)
            if ',' in lines[0]:
                return 'csv'
            
            # Check for other common CSV delimiters
            if '\t' in lines[0]:  # Tab-delimited
                return 'csv'
            if ';' in lines[0]:   # Semicolon-delimited
                return 'csv'
        
        return 'unknown'
        
    except Exception as e:
        print(f"Error detecting file format for {file_path}: {e}")
        return 'unknown'


def parse_csv_log_file(file_path):
    """
    Parse a CSV format DynamoDB metrics log file and extract timestamps and values.
    
    This function handles the specific CSV format used by the DynamoDB metrics collection
    tools. It supports two different data formats depending on the metric type.
    
    SUPPORTED FORMATS:
    ------------------
    1. Sample Count Format:
       DATAPOINTS <value> <timestamp> <unit>
       Example: DATAPOINTS 2935.0 2025-08-05T15:29:00+00:00 Milliseconds
    
    2. P99 Latency Format:
       DATAPOINTS <timestamp> <unit>
       EXTENDEDSTATISTICS <value>
       Example: 
       DATAPOINTS 2025-08-05T15:29:00+00:00 Milliseconds
       EXTENDEDSTATISTICS 19.360180913697853
    
    PARSING LOGIC:
    --------------
    1. File Type Detection:
       - Determines if file is sample_count or p99_latency based on path
       - Uses different parsing logic for each type
    
    2. Sample Count Parsing:
       - Looks for DATAPOINTS lines with 3+ parts
       - Extracts value (part 1) and timestamp (part 2)
       - Supports multiple timestamp formats
    
    3. P99 Latency Parsing:
       - Looks for DATAPOINTS lines with timestamp
       - Looks for EXTENDEDSTATISTICS lines with value
       - Pairs timestamp with value from next line
    
    4. Timestamp Parsing:
       - Supports multiple ISO 8601 formats
       - Handles timezone offsets and Z suffix
       - Falls back gracefully for unknown formats
    
    Args:
        file_path (str): Path to the individual log file
        
    Returns:
        list: List of tuples (timestamp, value) extracted from the file
        
    Example:
        >>> parse_csv_log_file('sample_count.log')
        [(datetime(2025, 8, 5, 15, 29), 2935.0), ...]
        
    Note:
        Returns empty list if file cannot be parsed or contains no valid data
    """
    timestamps_values = []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        lines = content.strip().split('\n')
        current_timestamp = None
        
        # Determine file type based on path
        is_sample_count = 'sample_count' in file_path or 'SampleCount' in file_path
        is_p99_latency = 'p99_latency' in file_path or 'P99' in file_path
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            
            # Look for DATAPOINTS pattern
            if line.startswith('DATAPOINTS'):
                parts = line.split('\t')  # Try tab separator first
                if len(parts) < 2:
                    parts = line.split()  # Fall back to space separator
                
                if is_sample_count and len(parts) >= 3:
                    # Sample count format: DATAPOINTS<tab>value<tab>timestamp<tab>unit
                    try:
                        value = float(parts[1])
                        timestamp_str = parts[2]
                        
                        # Parse timestamp
                        timestamp_formats = [
                            '%Y-%m-%dT%H:%M:%S+00:00',  # ISO format with timezone
                            '%Y-%m-%dT%H:%M:%SZ',       # ISO format
                            '%Y-%m-%d %H:%M:%S',        # Standard format
                            '%Y-%m-%d %H:%M',           # Short format
                        ]
                        
                        for fmt in timestamp_formats:
                            try:
                                timestamp = datetime.strptime(timestamp_str, fmt)
                                timestamps_values.append((timestamp, value))
                                break
                            except ValueError:
                                continue
                    except (ValueError, TypeError, IndexError):
                        pass
                
                elif is_p99_latency and len(parts) >= 2:
                    # P99 latency format: DATAPOINTS<tab>timestamp<tab>unit
                    # Value will be on next line with EXTENDEDSTATISTICS
                    timestamp_str = parts[1]
                    
                    # Parse timestamp
                    timestamp = None
                    timestamp_formats = [
                        '%Y-%m-%dT%H:%M:%S+00:00',  # ISO format with timezone
                        '%Y-%m-%dT%H:%M:%SZ',       # ISO format
                        '%Y-%m-%d %H:%M:%S',        # Standard format
                        '%Y-%m-%d %H:%M',           # Short format
                    ]
                    
                    for fmt in timestamp_formats:
                        try:
                            timestamp = datetime.strptime(timestamp_str, fmt)
                            current_timestamp = timestamp
                            break
                        except ValueError:
                            continue
                
            # Look for EXTENDEDSTATISTICS value (for p99_latency files)
            elif line.startswith('EXTENDEDSTATISTICS') and current_timestamp is not None:
                parts = line.split('\t')  # Try tab separator first
                if len(parts) < 2:
                    parts = line.split()  # Fall back to space separator
                    
                if len(parts) >= 2:
                    try:
                        value = float(parts[1])
                        timestamps_values.append((current_timestamp, value))
                        current_timestamp = None  # Reset for next pair
                    except (ValueError, TypeError):
                        pass
    
    except Exception as e:
        print(f"Error reading CSV file {file_path}: {e}")
        return []
    
    return timestamps_values


def parse_individual_log_file(file_path, file_format):
    """
    Parse an individual DynamoDB metrics log file and extract timestamps and values.
    Supports both JSON and CSV formats based on the specified format.
    
    Args:
        file_path (str): Path to the individual log file
        file_format (str): Format of the file ('json' or 'csv')
        
    Returns:
        list: List of tuples (timestamp, value) extracted from the file
    """
    if file_format == 'json':
        return parse_json_log_file(file_path)
    elif file_format == 'csv':
        return parse_csv_log_file(file_path)
    else:
        print(f"Warning: Unknown file format '{file_format}' for {file_path}")
        return []


def parse_json_log_file(file_path):
    """
    Parse a JSON format DynamoDB metrics log file and extract timestamps and values.
    
    Args:
        file_path (str): Path to the individual log file
        
    Returns:
        list: List of tuples (timestamp, value) extracted from the file
    """
    timestamps_values = []
    
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Parse the JSON content directly
        data = json.loads(content)
        
        # Extract datapoints from the JSON response
        if 'Datapoints' in data and isinstance(data['Datapoints'], list):
            for datapoint in data['Datapoints']:
                if 'Timestamp' not in datapoint:
                    continue
                
                timestamp_str = datapoint['Timestamp']
                value = None
                
                # Handle SampleCount (for sample_count files)
                if 'SampleCount' in datapoint and datapoint['SampleCount'] is not None:
                    value = datapoint['SampleCount']
                
                # Handle P99 latency (for p99_latency files)
                elif 'ExtendedStatistics' in datapoint and datapoint['ExtendedStatistics'] is not None:
                    if 'p99' in datapoint['ExtendedStatistics']:
                        value = datapoint['ExtendedStatistics']['p99']
                
                if value is not None:
                    # Parse timestamp (format: "2025-08-05T04:43:00Z")
                    try:
                        timestamp = datetime.strptime(timestamp_str, '%Y-%m-%dT%H:%M:%SZ')
                        timestamps_values.append((timestamp, float(value)))
                    except (ValueError, TypeError) as e:
                        continue
    
    except Exception as e:
        print(f"Error reading JSON file {file_path}: {e}")
        return []
    
    return timestamps_values


def find_all_regions(log_directory):
    """
    Find all regions in the log directory.
    
    Args:
        log_directory (str): Path to the directory containing log files
        
    Returns:
        list: List of region names found in the directory
    """
    regions = []
    
    # List all directories in the log directory
    for item in os.listdir(log_directory):
        item_path = os.path.join(log_directory, item)
        if os.path.isdir(item_path):
            # Check if it looks like a region directory by verifying it contains table directories
            # Don't rely on specific naming patterns, just check for subdirectories with expected structure
            if not item.startswith('.'):
                # Verify it contains subdirectories (tables) with operation subdirectories
                has_valid_structure = False
                try:
                    subdirs = [d for d in os.listdir(item_path) if os.path.isdir(os.path.join(item_path, d))]
                    # Check if at least one subdir has operation-like subdirectories
                    for subdir in subdirs[:3]:  # Check first few subdirs
                        subdir_path = os.path.join(item_path, subdir)
                        if os.path.isdir(subdir_path):
                            # Check for operation directories
                            op_dirs = [d for d in os.listdir(subdir_path) if os.path.isdir(os.path.join(subdir_path, d))]
                            # If we find directories that could be operations, this is likely a region
                            if op_dirs:
                                has_valid_structure = True
                                break
                except (OSError, PermissionError):
                    continue
                    
                if has_valid_structure:
                    regions.append(item)
    
    return sorted(regions)


def find_individual_log_files_for_region(log_directory, region):
    """
    Find all individual DynamoDB metrics log files for a specific region.
    
    IMPORTANT: This function only processes individual log files inside:
    - region/table/operation/p99_latency/*.log
    - region/table/operation/sample_count/*.log
    
    It explicitly EXCLUDES aggregate files like:
    - region/table/*_3hr.log
    - region/table/*_7day.log
    
    Args:
        log_directory (str): Path to the directory containing log files
        region (str): Region name to process
        
    Returns:
        dict: Dictionary with table names as keys and lists of file information as values
    """
    table_files = defaultdict(list)
    
    # Look for individual log files ONLY in p99_latency and sample_count directories
    # Structure: region/table/operation/metric_type/*.log
    # This pattern ensures we skip aggregate 3hr.log and 7day.log files at table level
    region_path = os.path.join(log_directory, region)
    pattern = os.path.join(region_path, "*", "*", "*", "*.log")
    
    for file_path in glob.glob(pattern, recursive=True):
        # Extract information from the path
        path_parts = file_path.split(os.sep)
        
        # Expected structure: log_dir/region/table/operation/metric_type/filename.log
        if len(path_parts) < 6:
            continue
            
        # Extract components from path
        table_name = path_parts[-4]
        operation = path_parts[-3]
        metric_type = path_parts[-2]
        filename = path_parts[-1]
        
        # CRITICAL: Only process files from p99_latency or sample_count directories
        # Skip any other files to avoid processing aggregate 3hr.log and 7day.log files
        if metric_type not in ['p99_latency', 'sample_count']:
            continue
        
        # Additional validation: Skip files that end with 3hr.log or 7day.log
        # (these shouldn't be in the metric_type directories, but this is defensive)
        if filename.endswith('3hr.log') or filename.endswith('7day.log'):
            print(f"  Warning: Skipping aggregate file found in {metric_type} directory: {filename}")
            continue
        
        # Determine period type from filename
        # Look for patterns like "20250805042954to20250805044954" (3hr) or "20250729073020to20250730073020" (7day)
        if 'to' in filename:
            # Extract the date range from filename
            date_range = filename.split('to')[0]
            
            # Determine if it's 3hr or 7day based on the date pattern
            # 3hr files have timestamps like "20250805042954" (same day)
            # 7day files have timestamps like "20250729073020" (different days)
            try:
                start_date = datetime.strptime(date_range, '%Y%m%d%H%M%S')
                end_part = filename.split('to')[1].replace('.log', '')
                end_date = datetime.strptime(end_part, '%Y%m%d%H%M%S')
                
                # Calculate time difference
                time_diff = end_date - start_date
                
                if time_diff.total_seconds() <= 4 * 3600:  # 4 hours or less
                    period_type = '3hr'
                else:
                    period_type = '7day'
                    
            except ValueError:
                # Fallback: if we can't parse, assume 3hr for recent files
                period_type = '3hr'
        else:
            period_type = '3hr'  # Default
        
        # Store file information
        table_files[table_name].append({
            'file_path': file_path,
            'operation': operation,
            'metric_type': metric_type,
            'period_type': period_type,
            'region': region
        })
    
    return table_files


def parse_table_detailed_log(log_directory, region_name):
    """
    Parse table_detailed.log file to extract table information.
    The file is expected to be at the top level of log_directory, not in region subdirectory.
    
    Args:
        log_directory (str): Path to the directory containing log files
        region_name (str): Region name to filter tables for
        
    Returns:
        dict: Dictionary with table names as keys and table details as values
    """
    table_details = {}
    
    # Look for table_detailed.log file at the top level directory first
    detailed_log_path = os.path.join(log_directory, "table_detailed.log")
    
    # If not found at top level, try in the region directory
    if not os.path.exists(detailed_log_path):
        detailed_log_path = os.path.join(log_directory, region_name, "table_detailed.log")
    
    if not os.path.exists(detailed_log_path):
        print(f"Warning: table_detailed.log file not found at {log_directory} or {os.path.join(log_directory, region_name)}")
        return table_details
    
    try:
        with open(detailed_log_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Parse the text content line by line
        current_table = None
        table_info = {}
        
        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # Check for table header with region
            if line.startswith('=== Table:'):
                # Check if this table is for the current region
                table_region_match = re.search(r'=== Table: ([\w-]+) \(Region: ([\w-]+)\)', line)
                if table_region_match:
                    table_name = table_region_match.group(1)
                    table_region = table_region_match.group(2)
                    
                    # Only process if it's for the current region
                    if table_region == region_name:
                        # Save previous table if exists
                        if current_table and table_info:
                            table_details[current_table] = table_info
                        
                        # Start new table
                        current_table = table_name
                        table_info = {
                            'TableName': current_table,
                            'TableSizeBytes': 'N/A',
                            'ItemCount': 'N/A',
                            'CreationDateTime': 'N/A',
                            'BillingModeSummary': 'N/A',
                            'ProvisionedThroughput': {},
                            'GlobalSecondaryIndexes': [],
                            'LocalSecondaryIndexes': [],
                            'StreamSpecification': {}
                        }
                    else:
                        # Skip tables from other regions
                        current_table = None
                        table_info = {}
                else:
                    # Fallback for table header without region
                    table_match = re.search(r'=== Table: ([\w-]+)', line)
                    if table_match:
                        current_table = table_match.group(1)
                        table_info = {
                            'TableName': current_table,
                            'TableSizeBytes': 'N/A',
                            'ItemCount': 'N/A',
                            'CreationDateTime': 'N/A',
                            'BillingModeSummary': 'N/A',
                            'ProvisionedThroughput': {},
                            'GlobalSecondaryIndexes': [],
                            'LocalSecondaryIndexes': [],
                            'StreamSpecification': {}
                        }
            
            # Parse table information from the ASCII table format
            elif current_table and line.startswith(f'Region: {region_name} |'):
                # Extract information from the line
                if 'TableStatus' in line:
                    # Extract status from format: ||  TableStatus              |  ACTIVE                               ||
                    status_match = re.search(r'TableStatus\s*\|\s*(\w+)', line)
                    if status_match:
                        table_info['TableStatus'] = status_match.group(1)
                elif 'CreationDateTime' in line:
                    # Extract creation date from format: ||  CreationDateTime         |  2024-09-23T15:40:28.722000-04:00     ||
                    date_match = re.search(r'CreationDateTime\s*\|\s*([\d\-T:.]+)', line)
                    if date_match:
                        table_info['CreationDateTime'] = date_match.group(1)
                elif 'ItemCount' in line:
                    # Extract item count from format: ||  ItemCount                |  1670604447                           ||
                    count_match = re.search(r'ItemCount\s*\|\s*(\d+)', line)
                    if count_match:
                        table_info['ItemCount'] = int(count_match.group(1))
                elif 'TableSizeBytes' in line:
                    # Extract size from format: ||  TableSizeBytes           |  1156168014599                        ||
                    size_match = re.search(r'TableSizeBytes\s*\|\s*(\d+)', line)
                    if size_match:
                        table_info['TableSizeBytes'] = int(size_match.group(1))
                elif 'BillingMode' in line:
                    # Extract billing mode from format: ||  BillingMode                                                   |  PAY_PER_REQUEST                                               ||
                    billing_match = re.search(r'BillingMode\s*\|\s*(\w+)', line)
                    if billing_match:
                        table_info['BillingModeSummary'] = billing_match.group(1)
                elif 'ReadCapacityUnits' in line:
                    # Extract read capacity from format: ||  ReadCapacityUnits                                                                                       |  0                 ||
                    read_match = re.search(r'ReadCapacityUnits\s*\|\s*(\d+)', line)
                    if read_match:
                        table_info['ProvisionedThroughput']['ReadCapacityUnits'] = int(read_match.group(1))
                elif 'WriteCapacityUnits' in line:
                    # Extract write capacity from format: ||  WriteCapacityUnits                                                                                      |  0                 ||
                    write_match = re.search(r'WriteCapacityUnits\s*\|\s*(\d+)', line)
                    if write_match:
                        table_info['ProvisionedThroughput']['WriteCapacityUnits'] = int(write_match.group(1))
                elif 'GlobalSecondaryIndexes' in line:
                    # Check if GSI exists (not None)
                    if 'None' not in line:
                        table_info['GlobalSecondaryIndexes'] = [{'IndexName': 'GSI'}]
                elif 'LocalSecondaryIndexes' in line:
                    # Check if LSI exists (not None)
                    if 'None' not in line:
                        table_info['LocalSecondaryIndexes'] = [{'IndexName': 'LSI'}]
                elif 'StreamSpecification' in line or 'StreamEnabled' in line:
                    # Check if streams are enabled
                    if 'true' in line.lower():
                        table_info['StreamSpecification']['StreamEnabled'] = True
        
        # Save the last table
        if current_table and table_info:
            table_details[current_table] = table_info
                
    except Exception as e:
        print(f"Error reading table_detailed.log file: {e}")
        return {}
    
    return table_details


def create_summary_worksheet(wb, table_details, max_sample_counts, max_p99_latencies):
    """
    Create a summary worksheet with table information and sorted max values.
    
    Args:
        wb: Excel workbook
        table_details (dict): Dictionary with table details
        max_sample_counts (list): Sorted list of (table_name, max_value, timestamp) for sample counts
        max_p99_latencies (list): Sorted list of (table_name, max_value, timestamp) for p99 latencies
    """
    if not table_details:
        print("No table details available for summary worksheet")
        return
    
    ws = wb.create_sheet(title="Table Summary")
    
    # Set up headers
    headers = [
        'Table Name',
        'Size (Bytes)',
        'Item Count',
        'Created Date',
        'Billing Type',
        'Provisioned Read Capacity',
        'Provisioned Write Capacity',
        'Secondary Indexes',
        'Streams Enabled'
    ]
    
    # Add headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add table data
    for row_idx, (table_name, details) in enumerate(table_details.items(), start=2):
        # Table Name
        ws.cell(row=row_idx, column=1, value=table_name)
        
        # Size
        size_bytes = details.get('TableSizeBytes', 'N/A')
        if size_bytes != 'N/A':
            # Convert to MB for readability
            size_mb = size_bytes / (1024 * 1024)
            ws.cell(row=row_idx, column=2, value=f"{size_mb:.2f} MB")
        else:
            ws.cell(row=row_idx, column=2, value=size_bytes)
        
        # Item Count
        ws.cell(row=row_idx, column=3, value=details.get('ItemCount', 'N/A'))
        
        # Created Date
        created_date = details.get('CreationDateTime', 'N/A')
        if created_date != 'N/A':
            try:
                # Parse ISO format date
                dt = datetime.fromisoformat(created_date.replace('Z', '+00:00'))
                ws.cell(row=row_idx, column=4, value=dt.strftime('%Y-%m-%d %H:%M:%S'))
            except:
                ws.cell(row=row_idx, column=4, value=created_date)
        else:
            ws.cell(row=row_idx, column=4, value=created_date)
        
        # Billing Type
        ws.cell(row=row_idx, column=5, value=details.get('BillingModeSummary', 'N/A'))
        
        # Provisioned Read Capacity
        provisioned = details.get('ProvisionedThroughput', {})
        read_capacity = provisioned.get('ReadCapacityUnits', 'N/A')
        ws.cell(row=row_idx, column=6, value=read_capacity)
        
        # Provisioned Write Capacity
        write_capacity = provisioned.get('WriteCapacityUnits', 'N/A')
        ws.cell(row=row_idx, column=7, value=write_capacity)
        
        # Secondary Indexes
        gsi_count = len(details.get('GlobalSecondaryIndexes', []))
        lsi_count = len(details.get('LocalSecondaryIndexes', []))
        total_indexes = gsi_count + lsi_count
        ws.cell(row=row_idx, column=8, value=f"{total_indexes} (GSI: {gsi_count}, LSI: {lsi_count})")
        
        # Streams Enabled
        stream_spec = details.get('StreamSpecification', {})
        streams_enabled = stream_spec.get('StreamEnabled', False)
        ws.cell(row=row_idx, column=9, value="Yes" if streams_enabled else "No")
    
    # Add empty row after table details
    empty_row = len(table_details) + 3
    ws[f'A{empty_row}'] = ""
    
    # Add Max Sample Counts section
    sample_count_header_row = empty_row + 1
    ws[f'A{sample_count_header_row}'] = "Max Sample Counts (Sorted by Value)"
    ws[f'A{sample_count_header_row}'].font = Font(bold=True, size=12)
    ws[f'A{sample_count_header_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Add headers for max sample counts
    sample_headers = ['Rank', 'Table Name', 'Max Sample Count', 'Timestamp']
    for col_idx, header in enumerate(sample_headers, start=1):
        cell = ws.cell(row=sample_count_header_row + 1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add max sample count data
    for rank, (table_name, max_value, timestamp) in enumerate(max_sample_counts, start=1):
        row_idx = sample_count_header_row + 1 + rank
        ws.cell(row=row_idx, column=1, value=rank)
        ws.cell(row=row_idx, column=2, value=table_name)
        ws.cell(row=row_idx, column=3, value=max_value)
        ws.cell(row=row_idx, column=4, value=timestamp.strftime('%Y-%m-%d %H:%M'))
        
        # Highlight the cell with max value
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Add empty row after sample counts
    p99_header_row = sample_count_header_row + 2 + len(max_sample_counts)
    ws[f'A{p99_header_row}'] = ""
    
    # Add Max P99 Latencies section
    ws[f'A{p99_header_row + 1}'] = "Max P99 Latencies (Sorted by Value)"
    ws[f'A{p99_header_row + 1}'].font = Font(bold=True, size=12)
    ws[f'A{p99_header_row + 1}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    # Add headers for max p99 latencies
    p99_headers = ['Rank', 'Table Name', 'Max P99 Latency', 'Timestamp']
    for col_idx, header in enumerate(p99_headers, start=1):
        cell = ws.cell(row=p99_header_row + 2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add max p99 latency data
    for rank, (table_name, max_value, timestamp) in enumerate(max_p99_latencies, start=1):
        row_idx = p99_header_row + 2 + rank
        ws.cell(row=row_idx, column=1, value=rank)
        ws.cell(row=row_idx, column=2, value=table_name)
        ws.cell(row=row_idx, column=3, value=max_value)
        ws.cell(row=row_idx, column=4, value=timestamp.strftime('%Y-%m-%d %H:%M'))
        
        # Highlight the cell with max value
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_excel_workbook(table_files, table_details, output_file, file_format, logger=None):
    """
    Create Excel workbook with data from individual log files and table summary.
    
    Args:
        table_files (dict): Dictionary with table names and their file information
        table_details (dict): Dictionary with table details for summary
        output_file (str): Path to the output Excel file
        file_format (str): Format of the log files ('json' or 'csv')
        logger: Logger instance for logging missing data
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Dynamically determine operations from the actual files found
    all_operations = set()
    for table_name, files in table_files.items():
        for file_info in files:
            all_operations.add(file_info['operation'])
    
    # Sort operations for consistent ordering
    operations = sorted(list(all_operations))
    
    # Lists to store max values for sorting
    max_sample_counts = []
    max_p99_latencies = []
    
    for table_name, files in table_files.items():
        print(f"Processing table: {table_name}")
        
        # Create single worksheet per table
        sheet_name = table_name
        if len(sheet_name) > 31:  # Excel sheet name limit
            sheet_name = table_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Set up headers
        ws['A1'] = table_name
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add section headers
        ws['A2'] = "Sample Count"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add operation names for Sample Count (rows 3-9)
        for i, operation in enumerate(operations, start=3):
            ws[f'A{i}'] = operation
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Empty row after Sample Count (row 10)
        ws['A10'] = ""
        
        # P99 Latency section header (row 11)
        ws['A11'] = "P99 Latency"
        ws['A11'].font = Font(bold=True, size=12)
        ws['A11'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add operation names for P99 Latency (rows 12-18)
        for i, operation in enumerate(operations, start=12):
            ws[f'A{i}'] = operation
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Empty row after P99 Latency (row 19)
        ws['A19'] = ""
        
        # Process files for this table
        sample_count_data = defaultdict(list)
        p99_latency_data = defaultdict(list)
        
        for file_info in files:
            operation = file_info['operation']
            metric_type = file_info['metric_type']
            period_type = file_info['period_type']
            
            # Parse the individual log file with specified format
            timestamps_values = parse_individual_log_file(file_info['file_path'], file_format)
            
            if timestamps_values:
                if metric_type == 'sample_count':
                    sample_count_data[operation].extend(timestamps_values)
                elif metric_type == 'p99_latency':
                    p99_latency_data[operation].extend(timestamps_values)
        
        # Sort data by timestamp for each operation
        for operation in operations:
            if operation in sample_count_data:
                sample_count_data[operation].sort(key=lambda x: x[0])
            if operation in p99_latency_data:
                p99_latency_data[operation].sort(key=lambda x: x[0])
        
        # Check for missing data and report
        missing_sample_count = []
        missing_p99_latency = []
        
        for operation in operations:
            if operation not in sample_count_data or not sample_count_data[operation]:
                missing_sample_count.append(operation)
            if operation not in p99_latency_data or not p99_latency_data[operation]:
                missing_p99_latency.append(operation)
        
        if missing_sample_count:
            warning_msg = f"No sample count data found for operations: {', '.join(missing_sample_count)}"
            print(f"  Warning: {warning_msg}")
            if logger:
                logger.warning(f"Table {table_name}: {warning_msg}")
        if missing_p99_latency:
            warning_msg = f"No P99 latency data found for operations: {', '.join(missing_p99_latency)}"
            print(f"  Warning: {warning_msg}")
            if logger:
                logger.warning(f"Table {table_name}: {warning_msg}")
        
        # Collect all unique timestamps from both metric types
        all_timestamps = set()
        
        # Collect timestamps from sample count data
        for operation in operations:
            if operation in sample_count_data:
                for timestamp, value in sample_count_data[operation]:
                    all_timestamps.add(timestamp)
        
        # Collect timestamps from p99 latency data
        for operation in operations:
            if operation in p99_latency_data:
                for timestamp, value in p99_latency_data[operation]:
                    all_timestamps.add(timestamp)
        
        # Sort timestamps
        sorted_timestamps = sorted(list(all_timestamps))
        
        # Create timestamp headers (columns B onwards)
        for col_idx, timestamp in enumerate(sorted_timestamps, start=2):
            header_value = timestamp.strftime('%Y-%m-%d %H:%M')
            ws.cell(row=1, column=col_idx, value=header_value)
            ws.cell(row=1, column=col_idx).font = Font(bold=True)
            ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add Sample Count data (rows 3-9)
        for row_idx, operation in enumerate(operations, start=3):
            # Create lookup dictionary for quick access
            sample_lookup = {}
            if operation in sample_count_data:
                for timestamp, value in sample_count_data[operation]:
                    sample_lookup[timestamp] = value
            
            # Fill data for each timestamp column
            for col_idx, timestamp in enumerate(sorted_timestamps, start=2):
                if timestamp in sample_lookup:
                    ws.cell(row=row_idx, column=col_idx, value=sample_lookup[timestamp])
        
        # Add P99 Latency data (rows 12-18)
        for row_idx, operation in enumerate(operations, start=12):
            # Create lookup dictionary for quick access
            p99_lookup = {}
            if operation in p99_latency_data:
                for timestamp, value in p99_latency_data[operation]:
                    p99_lookup[timestamp] = value
            
            # Fill data for each timestamp column
            for col_idx, timestamp in enumerate(sorted_timestamps, start=2):
                if timestamp in p99_lookup:
                    ws.cell(row=row_idx, column=col_idx, value=p99_lookup[timestamp])
        
        # Add Sample Count Max Values section (rows 21-28)
        ws['A21'] = "Sample Count Max Values"
        ws['A21'].font = Font(bold=True, size=12)
        ws['A21'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add operation names for Sample Count Max (rows 22-28)
        for i, operation in enumerate(operations, start=22):
            ws[f'A{i}'] = operation
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Calculate and add max Sample Count values (rows 22-28)
        table_max_sample_count = None
        table_max_sample_timestamp = None
        
        for row_idx, operation in enumerate(operations, start=22):
            if operation in sample_count_data and sample_count_data[operation]:
                # Find max value and its timestamp
                max_value = max(sample_count_data[operation], key=lambda x: x[1])
                max_timestamp, max_val = max_value
                
                # Track overall table max
                if table_max_sample_count is None or max_val > table_max_sample_count:
                    table_max_sample_count = max_val
                    table_max_sample_timestamp = max_timestamp
                
                # Add max value in column B
                ws.cell(row=row_idx, column=2, value=max_val)
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Add timestamp in column C
                ws.cell(row=row_idx, column=3, value=max_timestamp.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_idx, column=3).font = Font(bold=True)
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add to sorted list if we have data
        if table_max_sample_count is not None:
            max_sample_counts.append((table_name, table_max_sample_count, table_max_sample_timestamp))
        
        # Empty row after Sample Count Max (row 29)
        ws['A29'] = ""
        
        # Add P99 Latency Max Values section (rows 30-37)
        ws['A30'] = "P99 Latency Max Values"
        ws['A30'].font = Font(bold=True, size=12)
        ws['A30'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # Add operation names for P99 Latency Max (rows 31-37)
        for i, operation in enumerate(operations, start=31):
            ws[f'A{i}'] = operation
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'A{i}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # Calculate and add max P99 Latency values (rows 31-37)
        table_max_p99_latency = None
        table_max_p99_timestamp = None
        
        for row_idx, operation in enumerate(operations, start=31):
            if operation in p99_latency_data and p99_latency_data[operation]:
                # Find max value and its timestamp
                max_value = max(p99_latency_data[operation], key=lambda x: x[1])
                max_timestamp, max_val = max_value
                
                # Track overall table max
                if table_max_p99_latency is None or max_val > table_max_p99_latency:
                    table_max_p99_latency = max_val
                    table_max_p99_timestamp = max_timestamp
                
                # Add max value in column B
                ws.cell(row=row_idx, column=2, value=max_val)
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
                ws.cell(row=row_idx, column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Add timestamp in column C
                ws.cell(row=row_idx, column=3, value=max_timestamp.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row_idx, column=3).font = Font(bold=True)
                ws.cell(row=row_idx, column=3).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add to sorted list if we have data
        if table_max_p99_latency is not None:
            max_p99_latencies.append((table_name, table_max_p99_latency, table_max_p99_timestamp))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Sort the max value lists
    max_sample_counts.sort(key=lambda x: x[1], reverse=True)  # Sort by value descending
    max_p99_latencies.sort(key=lambda x: x[1], reverse=True)  # Sort by value descending
    
    # Create summary worksheet first
    if table_details:
        create_summary_worksheet(wb, table_details, max_sample_counts, max_p99_latencies)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")


def main():
    """
    Main function to process DynamoDB log files and create Excel output.
    """
    # Set up argument parser
    parser = argparse.ArgumentParser(
        description='Process DynamoDB metrics log files and create Excel output',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python process_dynamodb_logs.py -t /path/to/logs -d /path/to/output -c customer_name -f json
  python process_dynamodb_logs.py -t /path/to/logs -d /path/to/output -c customer_name -f csv
        """
    )
    
    parser.add_argument(
        '-t', '--target',
        required=True,
        help='Source directory containing DynamoDB log files'
    )
    
    parser.add_argument(
        '-d', '--destination',
        required=True,
        help='Destination directory for output Excel file'
    )
    
    parser.add_argument(
        '-c', '--customer',
        required=True,
        help='Customer name prefix for the Excel file'
    )
    
    parser.add_argument(
        '-f', '--format',
        required=True,
        choices=['json', 'csv'],
        help='Data format of the log files (json or csv)'
    )
    
    # Parse arguments
    args = parser.parse_args()
    
    log_directory = args.target
    destination_directory = args.destination
    customer_name = args.customer
    file_format = args.format
    
    # Validate source directory
    if not os.path.exists(log_directory):
        print(f"Error: Source directory '{log_directory}' does not exist.")
        sys.exit(1)
    
    if not os.path.isdir(log_directory):
        print(f"Error: '{log_directory}' is not a directory.")
        sys.exit(1)
    
    # Validate destination directory
    if not os.path.exists(destination_directory):
        print(f"Creating destination directory: {destination_directory}")
        try:
            os.makedirs(destination_directory, exist_ok=True)
        except Exception as e:
            print(f"Error creating destination directory: {e}")
            sys.exit(1)
    
    if not os.path.isdir(destination_directory):
        print(f"Error: '{destination_directory}' is not a directory.")
        sys.exit(1)
    
    # Create log file in destination directory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_filename = f"{customer_name}_{timestamp}.log"
    log_file_path = os.path.join(destination_directory, log_filename)
    
    # Set up logging to both console and file
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file_path),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    logger = logging.getLogger(__name__)
    
    logger.info(f"Starting DynamoDB log processing")
    logger.info(f"Source directory: {log_directory}")
    logger.info(f"Destination directory: {destination_directory}")
    logger.info(f"Customer name: {customer_name}")
    logger.info(f"File format: {file_format}")
    logger.info(f"Log file: {log_file_path}")
    
    print(f"Scanning directory: {log_directory}")
    print(f"Using format: {file_format}")
    print(f"Log file: {log_file_path}")
    
    # Find all regions in the directory
    regions = find_all_regions(log_directory)
    
    if not regions:
        error_msg = "No region directories found in the log directory."
        logger.error(error_msg)
        print(error_msg)
        print("Expected region directories containing table subdirectories with operation folders.")
        sys.exit(1)
    
    logger.info(f"Found {len(regions)} region(s): {', '.join(regions)}")
    print(f"Found {len(regions)} region(s): {', '.join(regions)}")
    
    # Process each region separately
    output_files = []
    
    for region_name in regions:
        logger.info(f"Processing region: {region_name}")
        print(f"\n{'='*60}")
        print(f"Processing region: {region_name}")
        print(f"{'='*60}")
        
        # Find all individual log files for this region
        table_files = find_individual_log_files_for_region(log_directory, region_name)
        
        if not table_files:
            warning_msg = f"No individual log files found for region {region_name}"
            logger.warning(warning_msg)
            print(f"  {warning_msg}")
            print(f"  Expected pattern: {region_name}/table/operation/metric_type/*.log")
            continue
        
        logger.info(f"Found {len(table_files)} tables with individual log files for region {region_name}")
        print(f"  Found {len(table_files)} tables with individual log files:")
        for table_name, files in table_files.items():
            logger.info(f"  Table {table_name}: {len(files)} files")
            print(f"    {table_name}: {len(files)} files")
        
        # Parse table detailed log for this region
        table_details = parse_table_detailed_log(log_directory, region_name)
        if table_details:
            logger.info(f"Found table details for {len(table_details)} tables in region {region_name}")
            print(f"  Found table details for {len(table_details)} tables")
        
        # Create output filename with customer and region prefix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{customer_name}_{region_name}_dynamodb_metrics_{timestamp}.xlsx"
        output_file = os.path.join(destination_directory, output_filename)
        
        # Create Excel workbook for this region
        logger.info(f"Creating Excel workbook for region {region_name}: {output_file}")
        create_excel_workbook(table_files, table_details, output_file, file_format, logger)
        output_files.append(output_file)
        
        logger.info(f"Excel file created successfully: {output_file}")
        print(f"  Excel file created: {output_file}")
    
    # Log final summary
    logger.info(f"Processing complete! Created {len(output_files)} Excel file(s)")
    print(f"\n{'='*60}")
    print(f"Processing complete!")
    print(f"\nCreated {len(output_files)} Excel file(s):")
    for output_file in output_files:
        logger.info(f"  - {os.path.basename(output_file)}")
        print(f"  - {os.path.basename(output_file)}")
    
    logger.info(f"Log file saved to: {log_file_path}")
    print(f"\nLog file saved to: {log_file_path}")


if __name__ == "__main__":
    main()
